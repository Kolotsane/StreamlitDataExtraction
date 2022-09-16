import PyPDF2
import os
import re
import openpyxl
import xlsxwriter
import streamlit as st
from PIL import Image
import pandas as pd
# import matplotlib.pyplot as plt
# import seaborn as se
import numpy as np
import base64
from streamlit_option_menu import option_menu
from datetime import date as dt


@st.cache(allow_output_mutation=True)
def get_base64_of_bin_file(png_file):
    with open(png_file, "rb") as f:
        data = f.read()
    return base64.b64encode(data).decode()


def build_markup_for_logo(
        png_file,
        background_position="50% 10%",
        margin_top="10%",
        image_width="100%",
        image_height="",
):
    binary_string = get_base64_of_bin_file(png_file)
    return """
            <style>
                [data-testid="stSidebarNav"] {
                    background-image: url("data:image/png;base64,%s");
                    background-repeat: no-repeat;
                    background-position: %s;
                    margin-top: %s;
                    background-size: %s %s;
                }
            </style>
            """ % (
        binary_string,
        background_position,
        margin_top,
        image_width,
        image_height,
    )


def add_logo(png_file):
    logo_markup = build_markup_for_logo(png_file)
    st.markdown(
        logo_markup,
        unsafe_allow_html=True,
    )


add_logo("lg.png")


# with st.sidebar:
#     # images
#     img = Image.open("logo.png")
#     st.image(img, width=300)
#     # Text/Title
#     st.write("""
#     <style>
#     @import url('https://fonts.googleapis.com/css2?family=Fascinate');
#     html, body, [class*="css"]  {
#    font-family: 'Verdana', cursive;
#    background: white;
#     }
#     </style>
#     """, unsafe_allow_html=True)


selected = option_menu(
    menu_title=None,
    options=["Home", "About Us", "Services", "Contact Us", "Logout"],
    icons=["house", "book", "gear", "envelope", "key"],
    menu_icon="cast",
    default_index=0,
    orientation="horizontal"
)


# save uploaded file
def save_uploaded_file(pdf_files):
    with open(os.path.join("invoices", pdf_files.name), "wb") as f:
        f.write(pdf_files.getbuffer())
    return st.success("File saved")


# files
uploaded_files = st.file_uploader("", type=['pdf'], accept_multiple_files=False)
if uploaded_files is not None:
    # file_details = {"filename":uploaded_files.name, "filetype":uploaded_files.type}
    save_uploaded_file(uploaded_files)


def create_xlsx_file():
    # Workbook() takes one, non-optional, argument
    # which is the filename that we want to create.
    workbook = xlsxwriter.Workbook('Invoice_Information.xlsx')

    # The workbook object is then used to add new
    # worksheet via the add_worksheet() method.
    worksheet = workbook.add_worksheet()

    # Use the worksheet object to write
    # data via the write() method.
    worksheet.write('A1', 'Company Name')
    worksheet.write('B1', 'Invoice Number')
    worksheet.write('C1', 'Email')
    worksheet.write('D1', 'Date')
    worksheet.write('E1', 'Total')
    worksheet.write('F1', 'Processed Date')

    workbook.close()


def extract_insert_to_xlsx_file():
    # Opening the created excel file
    excelsheet = openpyxl.load_workbook('Invoice_Information.xlsx')
    sheet = excelsheet['Sheet1']

    # Extracting data from the multiple pdf files
    for file_name in os.listdir('invoices'):
        # st.write(file_name)

        load_pdf = open(r'C:\\Users\\KOLOTSANE\\PycharmProjects\\streamlit\\invoices\\' + file_name, 'rb')
        read_pdf = PyPDF2.PdfFileReader(load_pdf, strict=False)
        page_count = read_pdf.getNumPages()
        first_page = read_pdf.getPage(0)
        page_content = first_page.extractText()

        try:
            dtoday = dt.today().strftime('%Y-%m-%d')
            # print(page_content)
            # Finding the array of totals and storing them in variables
            total = re.findall(r'(?<!\S)(?:(?:cad|[$]|usd|R|M|P) ?[\d,.]+|[\d.,]+(?:cad|[$]|usd))(?!\S)', page_content)
            if total:
                total1 = total[len(total) - 1]
                # st.write(total1)
            else:
                total1 = ""

            # Finding the array of Invoice dates and storing them in variables
            invoice_date = re.findall(
                r'(?:\d{1,2}[-/th|st|nd|rd\s]*)?(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)?[a-z\s,.]*(?:\d{1,2}[-/th|st|nd|rd)\s,]*)+(?:\d{2,4})+',
                page_content)

            # # sorted(invoice_date, key=lambda x: datetime.datetime.strptime(x, '%d-|/| %m-|/| %Y'))
            for i in invoice_date:
                match = re.search(r"[0-9]{2}\/[0-9]{2}\/[0-9]{4}", i)
                match1 = re.search(r"[0-9]{2}\-[0-9]{2}\-[0-9]{4}", i)
                match2 = re.findall(
                    r'(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)+[a-z\s,./]*(?:\d{1,2}[-/th|st|nd|rd)\s,]*)+(?:\d{4})',
                    i)
                if match:
                    date = match.group()
                    # date1 = date.split('\n', 0)[0]
                    # st.write(date)
                elif match1:
                    date = match1.group()
                    # st.write(date)
                elif match2:
                    date = match2[0]
                    # st.write(date)
                # else:
                #     date = ""

            # Finding the array of emails and storing them in variables
            email = re.findall(r'([a-zA-Z0-9._-]+@[a-zA-Z0-9._-]+\.[a-zA-Z0-9._-]+)', page_content)
            if email:
                email1 = email[0]
                company_name = re.search(r'([\w\.-]+)@([\w\.-]+)', email1).group(2).split(".")[0]
                # st.write(email1)
                # st.write(company_name)
            else:
                email1 = ""
                company_name = ""

            # Finding the array of Invoice Numbers and storing them in variables

            invoice_no = re.findall(r'[a-zA-Z._-]{2,4}[0-9]{4,12}', page_content)
            # invoice_no = re.search(r'Invoice Number|Invoice #|Invoice No.|Invoice Number:|Invoice Number.|Invoice #.|Invoice No:(.*)', page_content).group()
            if invoice_no:
                invoice_no1 = invoice_no[len(invoice_no) - 1]
                print(invoice_no1)
            else:
                invoice_no1 = ""

            # Finding the array of company names and storing them in variables
            # company_name = re.findall(r'\b[A-Z]\w+(?:\.com?)?(?:[ -]+(?:&[ -]+)?[A-Z]\w+(?:\.com?)?){0,2}[,\s]+(?i:ltd|llc|inc|plc|co(?:rp)?|group|holding|gmbh)\b', page_content)
            # if company_name:
            #     print(company_name)
            # else:
            #     company_name = " "

            print('\n')
            # Storing data in excel file

            last_row_number = sheet.max_row
            print(last_row_number)

            sheet.cell(column=1, row=last_row_number + 1).value = company_name
            sheet.cell(column=2, row=last_row_number + 1).value = invoice_no1
            sheet.cell(column=3, row=last_row_number + 1).value = email1
            sheet.cell(column=4, row=last_row_number + 1).value = date
            sheet.cell(column=5, row=last_row_number + 1).value = total1
            sheet.cell(column=6, row=last_row_number + 1).value = dtoday

            # saving a file
            excelsheet.save('Invoice_Information.xlsx')

        except:
            st.write("Process failed")
            print('\n')


def read_from_excel():
    # read by default 1st sheet of an excel file
    df = pd.read_excel('Invoice_Information.xlsx')
    with st.expander("Raw Data"):
        st.table(df)
        df.dropna(inplace=True)
        df["Total"] = df["Total"].str.replace("$", "", regex=False).astype(float)
        df['Date'] = pd.to_datetime(df['Date'])
    with st.expander("Clean Data"):
        st.table(df)
    df.to_csv('Clean_Data.csv', index=False)
    df.to_excel('Clean_Data.xlsx', index=False)


# adding a button
if st.button('Process'):
    create_xlsx_file()
    extract_insert_to_xlsx_file()
    read_from_excel()

    # video
    # video_file = open("Analytics.mp4", "rb").read()
    # st.video(video_file)

else:
    st.write("Click Process Button")
